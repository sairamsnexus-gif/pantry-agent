import os
import sys
import re
import math
import openpyxl
import pandas as pd
from datetime import datetime, timezone
from dotenv import load_dotenv
from supabase import create_client

sys.stdout.reconfigure(encoding='utf-8')
load_dotenv('.env')
load_dotenv('.env.env')

SUPABASE_URL = os.environ.get('SUPABASE_URL')
SUPABASE_KEY = os.environ.get('SUPABASE_KEY')

if not SUPABASE_URL or not SUPABASE_KEY:
    print("Error: SUPABASE_URL or SUPABASE_KEY is missing from environment.")
    sys.exit(1)

client = create_client(SUPABASE_URL, SUPABASE_KEY)

def categorize_item(name: str) -> str:
    name_upper = name.upper()
    if any(w in name_upper for w in ['DHALL', 'DHAL', 'DAL', 'RICE', 'ATTA', 'MAIDA', 'RAVA', 'SOOJI', 'WHEAT', 'SUGAR', 'SALT', 'POHA', 'AVAL']):
        return 'Staples'
    if any(w in name_upper for w in ['OIL', 'GHEE', 'BUTTER', 'VANASPATI']):
        return 'Cooking Essentials'
    if any(w in name_upper for w in ['MUSTARD', 'JEERA', 'CUMIN', 'PEPPER', 'CHILLI', 'CHILLY', 'TURMERIC', 'CORIANDER', 'MASALA', 'FENUGREEK', 'CARDAMOM', 'CLOVE', 'CINNAMON', 'HING', 'ASAFOETIDA', 'GARLIC', 'GINGER', 'TAMARIND', 'POWDER', 'SOMPH', 'FENNEL']):
        return 'Spices'
    if any(w in name_upper for w in ['SOAP', 'SHAMPOO', 'SURF', 'RIN', 'VIM', 'DETTOL', 'HARPIC', 'COMFORT', 'LIQUID', 'DETERGENT', 'CLEANER', 'SCRUB', 'DISHWASH', 'FABRIC', 'AERIAL', 'TIDE']):
        return 'Cleaning & Household'
    if any(w in name_upper for w in ['BISCUIT', 'COOKIE', 'RUSK', 'CHIPS', 'NOODLE', 'MAGGI', 'SNACK', 'CHOCOLATE', 'PASTA', 'VERMICELLI', 'SEVAI']):
        return 'Snacks & Packaged'
    if any(w in name_upper for w in ['TEA', 'COFFEE', 'BOOST', 'HORLICKS', 'BOURNVITA', 'MILK', 'CURD', 'PANEER', 'JUICE', 'SQUASH', 'BEVERAGE']):
        return 'Beverages & Dairy'
    if any(w in name_upper for w in ['PASTE', 'BRUSH', 'COLGATE', 'PEPSODENT', 'LOTION', 'CREAM', 'FACE', 'HAIR', 'OIL', 'POWDER', 'PERFUME', 'DEO', 'SHAVING']):
        return 'Personal Care'
    return 'Other Groceries'

def extract_unit(name: str) -> str:
    name_upper = name.upper()
    if re.search(r'\b(KG|KGS|KILO)\b', name_upper) or ' 1KG' in name_upper or ' 2KG' in name_upper or ' 5KG' in name_upper or ' 10KG' in name_upper or '1 KG' in name_upper:
        return 'kg'
    if re.search(r'\b(GM|GMS|GRAM|GRAMS)\b', name_upper) or ' 100GM' in name_upper or ' 200GM' in name_upper or ' 500GM' in name_upper or ' 50GM' in name_upper:
        return 'g'
    if re.search(r'\b(L|LTR|LITRE|LITER|ML)\b', name_upper) or ' 1L' in name_upper or ' 500ML' in name_upper or ' 200ML' in name_upper:
        return 'L'
    if 'PKT' in name_upper or 'PACK' in name_upper or 'BOX' in name_upper or 'TIN' in name_upper or 'BOTTLE' in name_upper:
        return 'pack'
    return 'units'

def get_consumption_and_threshold(category: str, unit: str):
    if unit == 'kg':
        return 0.08, 0.5
    elif unit == 'g':
        return 5.0, 50.0
    elif unit == 'L':
        return 0.05, 0.5
    elif unit == 'pack':
        return 0.04, 1.0
    else:
        return 0.05, 1.0

def seed_database():
    excel_path = 'Source/Grocery_Bills/GraceWorld/Extracted_Bills_Complete_Details.xlsx'
    if not os.path.exists(excel_path):
        print(f"Excel file not found at {excel_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['All Consolidated Items']

    # Header is at row 4
    # Columns: [Bill ID, Store Name, Billed Date, Item Description, HSN / Code, Quantity, MRP (₹), Unit Rate (₹), Total Amount (₹)]
    rows = []
    for r in range(5, ws.max_row + 1):
        bill_id = ws.cell(row=r, column=2).value
        store_name = ws.cell(row=r, column=3).value
        billed_date = ws.cell(row=r, column=4).value
        item_desc = ws.cell(row=r, column=5).value
        hsn = ws.cell(row=r, column=6).value
        qty = ws.cell(row=r, column=7).value
        mrp = ws.cell(row=r, column=8).value
        unit_rate = ws.cell(row=r, column=9).value
        total_amt = ws.cell(row=r, column=10).value

        if not item_desc:
            continue

        item_desc = str(item_desc).strip()
        store_name = str(store_name).strip() if store_name else 'Grace World'
        
        try:
            qty_val = float(qty) if qty is not None else 1.0
        except (ValueError, TypeError):
            qty_val = 1.0
            
        try:
            unit_rate_val = float(unit_rate) if unit_rate is not None else 0.0
        except (ValueError, TypeError):
            unit_rate_val = 0.0

        try:
            total_amt_val = float(total_amt) if total_amt is not None else (qty_val * unit_rate_val)
        except (ValueError, TypeError):
            total_amt_val = qty_val * unit_rate_val

        rows.append({
            'bill_id': bill_id,
            'store_name': store_name,
            'billed_date': billed_date,
            'item_name': item_desc,
            'quantity': qty_val,
            'unit': extract_unit(item_desc),
            'unit_price': unit_rate_val,
            'total_price': total_amt_val,
            'category': categorize_item(item_desc)
        })

    print(f"Parsed {len(rows)} purchase records from Excel.")

    # 1. First fetch existing inventory to preserve or update
    existing_inv_res = client.table("inventory").select("item_name, id").execute()
    existing_items = {item['item_name'] for item in existing_inv_res.data}
    print(f"Existing items in inventory: {len(existing_items)}")

    # 2. Build unique items list for inventory
    unique_items = {}
    for r in rows:
        name = r['item_name']
        if name not in unique_items:
            burn, thresh = get_consumption_and_threshold(r['category'], r['unit'])
            unique_items[name] = {
                'item_name': name,
                'category': r['category'],
                'current_stock': r['quantity'],
                'unit': r['unit'],
                'daily_consumption': burn,
                'min_threshold': thresh,
                'last_restocked': datetime.now(timezone.utc).strftime('%Y-%m-%d')
            }
        else:
            unique_items[name]['current_stock'] += r['quantity']

    # Upsert/Insert items into inventory in batches
    items_to_insert = [v for k, v in unique_items.items() if k not in existing_items]
    items_to_update = [v for k, v in unique_items.items() if k in existing_items]

    print(f"Items to insert into inventory: {len(items_to_insert)}")
    print(f"Items already present in inventory: {len(items_to_update)}")

    batch_size = 50
    for i in range(0, len(items_to_insert), batch_size):
        batch = items_to_insert[i:i+batch_size]
        client.table("inventory").insert(batch).execute()
        print(f"Inserted inventory batch {i//batch_size + 1} ({len(batch)} items)")

    # 3. Clean and populate purchase_history table
    print("Clearing prior records in purchase_history (if any)...")
    try:
        client.table("purchase_history").delete().neq("id", 0).execute()
    except Exception as e:
        print("Note on clear:", e)

    # Insert 208 purchase rows into purchase_history
    purchase_batches = []
    for r in rows:
        purchase_batches.append({
            'item_name': r['item_name'],
            'store_name': r['store_name'],
            'quantity': r['quantity'],
            'unit': r['unit'],
            'unit_price': r['unit_price'],
            'total_price': r['total_price']
        })

    print(f"Inserting {len(purchase_batches)} rows into purchase_history...")
    for i in range(0, len(purchase_batches), batch_size):
        batch = purchase_batches[i:i+batch_size]
        client.table("purchase_history").insert(batch).execute()
        print(f"Inserted purchase_history batch {i//batch_size + 1} ({len(batch)} rows)")

    # Verify counts
    final_inv = client.table("inventory").select("id", count="exact").execute()
    final_ph = client.table("purchase_history").select("id", count="exact").execute()
    print("\n=== Seeding Verification ===")
    print(f"Total inventory items in Supabase: {final_inv.count if final_inv.count is not None else len(final_inv.data)}")
    print(f"Total purchase records in Supabase: {final_ph.count if final_ph.count is not None else len(final_ph.data)}")
    print("Baseline item stock and Grace baseline rates seeded successfully!")

if __name__ == '__main__':
    seed_database()
